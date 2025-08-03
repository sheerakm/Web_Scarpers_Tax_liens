# import requests
# from googlesearch import search  # pip install google
# from bs4 import BeautifulSoup    # pip install beautifulsoup4
#
# OLLAMA_URL = "http://localhost:11434/api/generate"
# MODEL_NAME = "llama3"  # Adjust if your model name differs
#
# def query_ollama(prompt):
#     response = requests.post(OLLAMA_URL, json={
#         "model": MODEL_NAME,
#         "prompt": prompt,
#         "stream": False
#     })
#     return response.json()["response"]
#
# def fetch_text(url):
#     try:
#         r = requests.get(url, timeout=5, headers={'User-Agent': 'Mozilla/5.0'})
#         soup = BeautifulSoup(r.text, 'html.parser')
#         for script in soup(["script", "style"]):
#             script.decompose()
#         return soup.get_text(separator=' ', strip=True)
#     except:
#         return ""
#
# def main(user_query):
#     urls = list(search(user_query, num=10, stop=10))
#     print(f"[+] Found {len(urls)} results for query: '{user_query}'\n")
#
#     for url in urls:
#         print(f"[URL] {url}")
#         content = fetch_text(url)
#         if not content:
#             print("[-] Failed to fetch content.\n")
#             continue
#
#         short_content = content[:3000]  # keep within context window
#         prompt = f"User is searching for: \"{user_query}\"\n\nDoes the following page content help? Answer yes or no, then explain shortly:\n\n{short_content}"
#         result = query_ollama(prompt)
#         print(f"[LLaMA 3] {result.strip()}\n{'-'*60}")
#
# if __name__ == "__main__":
#     query = "how to fine-tune llama 3 with custom data"
#     main(query)

import ollama
from googlesearch import search
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import requests

MODEL_NAME = "llama3"

def query_ollama(prompt):
    response = ollama.chat(model=MODEL_NAME, messages=[
        {"role": "user", "content": prompt}
    ])
    return response['message']['content']

def fetch_text(url):
    try:
        r = requests.get(url, timeout=5, headers={'User-Agent': 'Mozilla/5.0'})
        soup = BeautifulSoup(r.text, 'html.parser')
        for script in soup(["script", "style"]):
            script.decompose()
        return soup.get_text(separator=' ', strip=True)
    except:
        return ""



def get_best_answer(user_query, question,  example_answer, target_country):
    urls = list(search(user_query, num=5, stop=5))

    for url in urls:
        print("url is: ", url)
        if url.endswith(".pdf"):
            continue
        content = fetch_text(url)

        if not content or "not found" in content.lower() or len(content) < 500:
            continue

        if len(content) > 7000:
            content = content[:7000]

        prompt = (
            f"Google search was: '{user_query}'\n\n"
            f"Example answer style:\n{example_answer}\n\n"
            f"Now based on this website content (below), answer the following question **as briefly and factually as possible**, in the same format as the example. Do not apologize or explain anything extra.\n\n"
            f"Question: {question}\n"
            f"County: {target_country}\n"
            f"Website content:\n{content}"
        )

        raw_result = query_ollama(prompt).strip()

        if "Answer:" in raw_result:
            result = raw_result.split("Answer:", 1)[-1].strip()
        else:
            result = raw_result

        # print("answer, and questions are: ")

        print(question, result)

        return f"{{ {result}, {url} }}"

    return "No relevant info found."

def fill_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active

    questions = [ws.cell(row=1, column=j).value for j in range(2, ws.max_column + 1)]
    llm_guidance = [ws.cell(row=2, column=j).value for j in range(2, ws.max_column + 1)]
    example_answers = [ws.cell(row=3, column=j).value for j in range(2, ws.max_column + 1)]
    counties = [ws.cell(row=i, column=1).value for i in range(4, ws.max_row + 1)][0:50]

    for row_idx, county in enumerate(counties, start=4):
        for col_idx, (question, context, example) in enumerate(zip(questions, llm_guidance, example_answers), start=2):
            query = f"{county} county {question}"
            # print(f"[+] Query: {query}")
            answer = get_best_answer(query, context, example, county)
            ws.cell(row=row_idx, column=col_idx).value = answer
            # print(f"[✓] Filled {county} - {question}\n")

    wb.save("filled_" + filename)


if __name__ == "__main__":
    fill_excel("Top 50 Counties - Copy.xlsx")
